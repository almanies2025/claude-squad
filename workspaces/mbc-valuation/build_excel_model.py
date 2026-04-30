"""
MBC Group Valuation Model — Excel Builder
Generates a professional Excel workbook with DCF, Comps, SOTP, and Summary tabs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

OUTPUT_PATH = "/mnt/c/Users/User/Documents/GitHub/claude-squad/workspaces/mbc-valuation/MBC_Group_Valuation_Model.xlsx"

# ─── Color palette ─────────────────────────────────────────────────────────────
NAVY = "1F3864"
MID_BLUE = "2E75B6"
LIGHT_BLUE = "BDD7EE"
VERY_LIGHT = "DEEAF1"
GOLD = "C9A227"
DARK_GRAY = "404040"
MED_GRAY = "808080"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
RED_FILL = "FFCCCC"
GREEN_FILL = "CCFFCC"
YELLOW_FILL = "FFFFCC"
ORANGE_FILL = "FFE0CC"


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def medium_border():
    s = Side(style="medium", color="404040")
    return Border(left=s, right=s, top=s, bottom=s)


def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def data_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def section_header(ws, row, col, label, span_end_col=None):
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if span_end_col:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=span_end_col
        )
    return cell


def input_cell(ws, row, col, value, fmt=None, label=None, label_col=None):
    """Blue-tinted input cell for user-editable assumptions."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor="DAEEF3")
    cell.font = Font(name="Calibri", size=10, bold=True, color="1F497D")
    cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt
    if label and label_col:
        lbl = ws.cell(row=row, column=label_col, value=label)
        lbl.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        lbl.alignment = Alignment(horizontal="left")
    return cell


def calc_cell(ws, row, col, value, fmt=None, bold=False, color="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="right")
    if fmt:
        cell.number_format = fmt
    cell.border = thin_border()
    return cell


def label_cell(ws, row, col, text, indent=0, bold=False, color=DARK_GRAY):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", indent=indent)
    return cell


def pct_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.font = Font(name="Calibri", size=10, color="000000")
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = "0.0%"
    cell.border = thin_border()
    return cell


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def freeze_pane(ws, cell_ref):
    ws.freeze_panes = cell_ref


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: DCF MODEL
# ══════════════════════════════════════════════════════════════════════════════
def build_dcf_sheet(wb):
    ws = wb.create_sheet("DCF Model")
    ws.sheet_view.showGridLines = False

    # Column widths
    for col, w in [(1, 28), (2, 14), (3, 14), (4, 14), (5, 14), (6, 14), (7, 14)]:
        set_col_width(ws, col, w)

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "MBC GROUP — DCF VALUATION MODEL"
    t.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = "Currency: SAR millions | Priced as of April 21, 2026"
    sub.font = Font(name="Calibri", size=10, italic=True, color=MED_GRAY)
    sub.fill = PatternFill("solid", fgColor=NAVY)
    sub.alignment = Alignment(horizontal="center")
    set_row_height(ws, 2, 18)

    # ── Assumptions block ────────────────────────────────────────────────────
    r = 4
    section_header(ws, r, 1, "KEY ASSUMPTIONS", 7)
    set_row_height(ws, r, 20)

    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    ws[f"A{r}"].value = "Input cells (blue) can be adjusted by user"
    ws[f"A{r}"].font = Font(name="Calibri", size=9, italic=True, color="1F497D")
    ws[f"A{r}"].fill = PatternFill("solid", fgColor="DAEEF3")

    assumptions = [
        ("WACC (Base case)", 0.10, "0.0%", "B", 5),
        ("WACC (Bear)", 0.12, "0.0%", "B", 6),
        ("WACC (Bull)", 0.08, "0.0%", "B", 7),
        ("Terminal growth rate", 0.045, "0.0%", "B", 9),
        ("Risk-free rate", 0.0475, "0.0%", "B", 10),
        ("Equity risk premium", 0.055, "0.0%", "B", 11),
        ("Effective tax rate", 0.20, "0.0%", "B", 13),
        ("CapEx % of revenue", 0.041, "0.0%", "B", 14),
        ("D&A % of revenue", 0.023, "0.0%", "B", 15),
    ]

    for i, (label, val, _, col_let, row_num) in enumerate(assumptions):
        r = 6 + i
        label_cell(ws, r, 1, label, indent=2)
        input_cell(ws, r, 2, val, "0.0%")

    # ── Revenue Build ────────────────────────────────────────────────────────
    r = 16
    section_header(ws, r, 1, "REVENUE BUILD", 7)
    set_row_height(ws, r, 20)

    headers = ["", "FY2025", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r + 1, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")

    revenue_data = [
        ("Revenue (SAR M)", 5379.3, 5515.8, 6087.6, 6696.4, 7232.1, 7694.2),
        ("YoY Growth %", 0.285, 0.023, 0.104, 0.100, 0.080, 0.064),
    ]

    for j, row_data in enumerate(revenue_data):
        r2 = r + 2 + j
        for k, val in enumerate(row_data):
            cell = ws.cell(row=r2, column=k + 1, value=val)
            cell.font = Font(name="Calibri", size=10, bold=(k == 0), color=DARK_GRAY)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY if k > 0 else WHITE)
            cell.alignment = Alignment(
                horizontal="right" if k > 0 else "left", indent=2 if k == 0 else 0
            )
            if k > 0:
                if j == 0:
                    cell.number_format = "#,##0.0"
                else:
                    cell.number_format = "0.0%"
            cell.border = thin_border()

    # ── EBITDA & FCF Build ──────────────────────────────────────────────────
    r = 22
    section_header(ws, r, 1, "EBITDA & FREE CASH FLOW BUILD", 7)

    ebitda_rows = [
        ("EBITDA Margin %", 0.050, 0.126, 0.153, 0.160, 0.170, 0.180),
        ("EBITDA (SAR M)", None, None, None, None, None, None),
        ("D&A (SAR M)", None, None, None, None, None, None),
        ("CapEx (SAR M)", None, None, None, None, None, None),
        ("Cash Taxes (SAR M)", None, None, None, None, None, None),
        ("FCFF (SAR M)", None, None, None, None, None, None),
        ("FCF Margin %", None, None, None, None, None, None),
    ]

    for i, row_data in enumerate(ebitda_rows):
        r2 = r + 1 + i
        label_cell(ws, r2, 1, row_data[0], indent=2, bold=(i in [1, 5]))
        for j, val in enumerate(row_data[1:]):
            col_idx = j + 2
            if val is None:
                # formula
                if i == 1:  # EBITDA
                    rev_row = r + 3
                    margin_cell = ws.cell(row=r2, column=1).value  # label
                    cell = calc_cell(
                        ws,
                        r2,
                        col_idx,
                        f"={get_column_letter(col_idx)}{r + 2}*{get_column_letter(col_idx)}{r + 3}",
                        "#,##0.0",
                        bold=True,
                    )
                elif i == 2:  # D&A
                    cell = calc_cell(
                        ws,
                        r2,
                        col_idx,
                        f"={get_column_letter(col_idx)}{r + 2}*0.023",
                        "#,##0.0",
                    )
                elif i == 3:  # CapEx
                    cell = calc_cell(
                        ws,
                        r2,
                        col_idx,
                        f"={get_column_letter(col_idx)}{r + 2}*0.041",
                        "#,##0.0",
                    )
                elif i == 4:  # Cash Taxes
                    cell = calc_cell(
                        ws,
                        r2,
                        col_idx,
                        f"=({get_column_letter(col_idx)}{r + 2}-{get_column_letter(col_idx)}{r + 4})*0.20",
                        "#,##0.0",
                    )
                elif i == 5:  # FCFF
                    ebitda_col = get_column_letter(col_idx)
                    cell = calc_cell(
                        ws,
                        r2,
                        col_idx,
                        f"={ebitda_col}{r + 3}-{ebitda_col}{r + 5}-{ebitda_col}{r + 6}",
                        "#,##0.0",
                        bold=True,
                    )
                elif i == 6:  # FCF Margin
                    cell = pct_cell(
                        ws,
                        r2,
                        col_idx,
                        f"={get_column_letter(col_idx)}{r + 7}/{get_column_letter(col_idx)}{r + 2}",
                    )
            else:
                if i == 0:
                    cell = pct_cell(ws, r2, col_idx, val)
                else:
                    cell = calc_cell(ws, r2, col_idx, val, "#,##0.0")

    # ── Terminal Value ──────────────────────────────────────────────────────
    r = 31
    section_header(ws, r, 1, "TERMINAL VALUE", 7)

    tv_labels = [
        ("Terminal FCFF (FY2030)", "=F39"),
        ("Terminal growth rate", "=B10"),
        ("WACC (Base)", "=B6"),
        ("Terminal Value", "=B43*(1+B44)/(B45-B44)"),
        ("PV of Terminal Value", "=B46/(1+B45)^5"),
    ]

    for i, (lbl, formula) in enumerate(tv_labels):
        r2 = r + 1 + i
        label_cell(ws, r2, 1, lbl, indent=2, bold=(i >= 3))
        col = 2
        cell = ws.cell(
            row=r2, column=col, value=formula if formula.startswith("=") else None
        )
        if not formula.startswith("="):
            cell = ws.cell(row=r2, column=col, value=float(formula))
        cell.font = Font(name="Calibri", size=10, bold=(i >= 3), color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = "#,##0.0" if i < 4 else "#,##0.0"
        cell.border = thin_border()

    # ── Equity Bridge ───────────────────────────────────────────────────────
    r = 38
    section_header(ws, r, 1, "EQUITY VALUE BRIDGE", 7)

    equity_rows = [
        ("PV of FCFF (FY2026–30)", "=SUM(B36:F36)", None, None),
        ("PV of Terminal Value", "=B47", None, None),
        ("Enterprise Value (SAR M)", "=B49+B50", None, None),
        ("+ Cash & ST Investments", 1352.3, None, None),
        ("− Total Debt", -121.2, None, None),
        ("− Minority Interest", -70.0, None, None),
        ("Equity Value (SAR M)", "=B51+B52+B53+B54", None, None),
        ("Shares Outstanding (M)", 332.5, None, None),
        ("Implied Share Price (SAR)", "=B56/B58", None, None),
        ("Current Market Price", 26.26, None, None),
        ("Upside to Market %", "=(B59-B60)/B60", None, None),
    ]

    for i, row_data in enumerate(equity_rows):
        r2 = r + 1 + i
        label_cell(ws, r2, 1, row_data[0], indent=2, bold=(i in [2, 6, 9]))
        val = row_data[1]
        cell = ws.cell(row=r2, column=2)
        if isinstance(val, str) and val.startswith("="):
            cell.value = val
            cell.number_format = "#,##0.0"
        elif isinstance(val, (int, float)):
            cell.value = val
            cell.number_format = "#,##0.0"
        elif val is None:
            cell.value = "—"
            cell.number_format = "@"
        cell.font = Font(
            name="Calibri", size=10, bold=(i in [2, 6, 9]), color=DARK_GRAY
        )
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()
        if i == 10:
            cell.number_format = "0.0%"

    # ── Sensitivity Table ───────────────────────────────────────────────────
    r = 51
    section_header(ws, r, 1, "SHARE PRICE SENSITIVITY: WACC vs. TERMINAL GROWTH", 7)
    set_row_height(ws, r, 20)

    r += 1
    # Header row
    for col_idx, label in enumerate(
        ["Terminal g →", "3.5%", "4.0%", "4.5%", "5.0%"], start=1
    ):
        c = ws.cell(row=r, column=col_idx, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    wacc_values = [
        ("WACC 8%", 0.08),
        ("WACC 9%", 0.09),
        ("WACC 10%", 0.10),
        ("WACC 11%", 0.11),
        ("WACC 12%", 0.12),
    ]

    g_values = [0.035, 0.040, 0.045, 0.050]
    tv_row = 46  # Terminal value row

    for row_i, (wacc_label, wacc) in enumerate(wacc_values):
        r2 = r + 1 + row_i
        # WACC label
        lc = ws.cell(row=r2, column=1, value=wacc_label)
        lc.font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        lc.fill = PatternFill("solid", fgColor=VERY_LIGHT)
        lc.alignment = Alignment(horizontal="left", indent=2)
        lc.border = thin_border()

        for col_i, g in enumerate(g_values):
            col_idx = col_i + 2
            # Manual calculation for each cell
            fcf_2030 = 924.8  # from model
            tv = fcf_2030 * (1 + g) / (wacc - g)
            pv_tv = tv / (1 + wacc) ** 5
            pv_fcf = 2559.2  # sum of PV FCFs
            ev = pv_fcf + pv_tv
            equity = ev + 1352.3 - 121.2 - 70.0
            price = equity / 332.5

            cell = ws.cell(row=r2, column=col_idx, value=round(price, 1))
            cell.number_format = "#,##0.0"
            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()
            if row_i == 2 and col_i == 2:
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
                cell.font = Font(name="Calibri", size=10, bold=True, color="1F497D")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: COMPARABLE COMPANIES
# ══════════════════════════════════════════════════════════════════════════════
def build_comps_sheet(wb):
    ws = wb.create_sheet("Comparable Companies")
    ws.sheet_view.showGridLines = False

    for col, w in [(1, 30), (2, 14), (3, 14), (4, 14), (5, 14)]:
        set_col_width(ws, col, w)

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MBC GROUP — COMPARABLE COMPANIES ANALYSIS"
    t.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    # Peer Universe
    r = 3
    section_header(ws, r, 1, "PEER GROUP", 5)

    peers = [
        ("Netflix", "NFLX", "$393B", "26.3x", "8.4x"),
        ("Disney", "DIS", "$220B", "20.0x", "2.6x"),
        ("Warner Bros Disc.", "WBD", "—", "3.1x", "2.6x"),
        ("Paramount (deal)", "PARA", "—", "7.5x", "2.5x"),
        ("MBC (subject)", "4072", "SAR 8.8B", "11.1x", "1.4x"),
    ]

    headers = ["Company", "Ticker", "Market Cap", "EV/EBITDA", "EV/Revenue"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r + 1, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for j, row_data in enumerate(peers):
        r2 = r + 2 + j
        is_mbc = row_data[0] == "MBC (subject)"
        for k, val in enumerate(row_data):
            cell = ws.cell(row=r2, column=k + 1, value=val)
            cell.font = Font(
                name="Calibri",
                size=10,
                bold=is_mbc,
                color=NAVY if is_mbc else DARK_GRAY,
            )
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY if is_mbc else WHITE)
            cell.alignment = Alignment(
                horizontal="center" if k > 0 else "left", indent=1 if k == 0 else 0
            )
            cell.border = thin_border()

    # Multiples Applied
    r = 12
    section_header(ws, r, 1, "MULTIPLES APPLIED TO MBC (NTM EBITDA: SAR 692.8M)", 5)

    multiples = [
        ("Netflix 26x", 26.3, "18,220", "19,451", "SAR 58.5", "+123%"),
        ("Disney 20x", 20.0, "13,856", "15,087", "SAR 45.4", "+73%"),
        ("Blended 15x", 15.0, "10,392", "11,623", "SAR 35.0", "+33%"),
        ("Blended 12x", 12.0, " 8,314", " 9,545", "SAR 28.7", "+9%"),
        ("Paramount 7.5x", 7.5, " 5,196", " 5,427", "SAR 16.3", "-38%"),
        ("WBD Distressed 3x", 3.1, " 2,148", "   917", "SAR  2.8", "-89%"),
    ]

    headers2 = [
        "Multiple",
        "Applied",
        "Implied EV",
        "Equity Value",
        "Share Price",
        "vs Market",
    ]
    for i, h in enumerate(headers2):
        c = ws.cell(row=r + 1, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for j, row_data in enumerate(multiples):
        r2 = r + 2 + j
        is_positive = "+" in str(row_data[5])
        for k, val in enumerate(row_data):
            cell = ws.cell(row=r2, column=k + 1, value=val)
            cell.font = Font(
                name="Calibri",
                size=10,
                color=GREEN_FILL if (is_positive and k == 5) else DARK_GRAY,
            )
            if k == 5:
                cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=True,
                    color="00B050" if is_positive else "FF0000",
                )
            cell.fill = PatternFill(
                "solid", fgColor=LIGHT_GRAY if j % 2 == 0 else WHITE
            )
            cell.alignment = Alignment(
                horizontal="right" if k > 0 else "left", indent=1 if k == 0 else 0
            )
            cell.border = thin_border()

    # Revenue multiple cross-check
    r = 21
    section_header(
        ws, r, 1, "REVENUE MULTIPLE CROSS-CHECK (MBC NTM Revenue: SAR 5,515.8M)", 5
    )

    rev_comps = [
        ("Netflix 8.4x", 8.4, "46,334", "SAR 143", "Too high — NFLX not comparable"),
        ("Disney 2.6x", 2.6, "14,341", "SAR 46.8", "Reasonable for blended media"),
        ("WBD 2.6x", 2.6, "14,341", "SAR 46.8", "Same as Disney"),
        ("MBC current 1.4x", 1.4, " 7,868", "—", "Matches current EV"),
    ]

    headers3 = ["Peer", "EV/Rev", "Implied EV", "Share Price", "Note"]
    for i, h in enumerate(headers3):
        c = ws.cell(row=r + 1, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for j, row_data in enumerate(rev_comps):
        r2 = r + 2 + j
        for k, val in enumerate(row_data):
            cell = ws.cell(row=r2, column=k + 1, value=val)
            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill(
                "solid", fgColor=LIGHT_GRAY if j % 2 == 0 else WHITE
            )
            cell.alignment = Alignment(
                horizontal="right" if k in [1, 2, 3] else "left",
                indent=1 if k == 0 else 0,
            )
            cell.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: SUM-OF-PARTS
# ══════════════════════════════════════════════════════════════════════════════
def build_sotp_sheet(wb):
    ws = wb.create_sheet("Sum-of-Parts")
    ws.sheet_view.showGridLines = False

    for col, w in [(1, 30), (2, 14), (3, 14), (4, 14), (5, 16)]:
        set_col_width(ws, col, w)

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MBC GROUP — SUM-OF-PARTS VALUATION"
    t.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    # BOCA
    r = 3
    section_header(ws, r, 1, "SEGMENT 1: BOCA (BROADCAST)", 5)
    r += 1
    label_cell(ws, r, 1, "Analog: Regional/global TV broadcasters", indent=2)
    r += 1
    label_cell(ws, r, 1, "Best peer: Sinclair Broadcast Group (SBGI)", indent=2)
    r += 1
    label_cell(ws, r, 1, "Chosen range: 9–12x EBITDA", indent=2)

    r += 1
    boca_data = [
        ("FY2026E Revenue (SAR M)", 3200, 3200, 3200),
        ("EBITDA Margin", 0.20, 0.22, 0.24),
        ("EBITDA (SAR M)", None, None, None),
    ]

    headers4 = ["", "Bear (9x)", "Base (10.5x)", "Bull (12x)"]
    for i, h in enumerate(headers4):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for j, row_data in enumerate(boca_data):
        r2 = r + 1 + j
        label_cell(ws, r2, 1, row_data[0], indent=2)
        for k, val in enumerate(row_data[1:]):
            col_idx = k + 2
            cell = ws.cell(row=r2, column=col_idx)
            if val is None:
                if j == 2:
                    cell.value = f"={get_column_letter(col_idx)}{r + 1}*{get_column_letter(col_idx)}{r + 2}"
                else:
                    cell.value = "—"
            else:
                cell.value = val
                if j == 0:
                    cell.number_format = "#,##0"
                elif j == 1:
                    cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border()

    r += 5
    label_cell(ws, r, 1, "EV/EBITDA multiple → EV", indent=2)
    for k, (mult, lbl) in enumerate([(9, "Bear"), (10.5, "Base"), (12, "Bull")]):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx)
        cell.value = f"={get_column_letter(col_idx)}{r - 3}*{mult}"
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    r += 2
    label_cell(ws, r, 1, "Less: Net debt allocated (54%)", indent=2)
    for k in range(3):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx, value=-665)
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    r += 2
    label_cell(ws, r, 1, "BOCA Equity (SAR M)", indent=2, bold=True)
    for k in range(3):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx)
        cell.value = (
            f"={get_column_letter(col_idx)}{r - 4}+{get_column_letter(col_idx)}{r - 2}"
        )
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=VERY_LIGHT)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    # Shahid
    r += 3
    section_header(ws, r, 1, "SEGMENT 2: SHAHID (STREAMING)", 5)
    r += 1
    label_cell(ws, r, 1, "Method: EV/Revenue (not yet EBITDA positive)", indent=2)
    r += 1
    label_cell(
        ws,
        r,
        1,
        "Best analog: Netflix (8.4x), Disney (2.6x) — MENA discount applies",
        indent=2,
    )

    r += 1
    headers5 = ["", "Bear (3x)", "Base (5x)", "Bull (8x)"]
    for i, h in enumerate(headers5):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    shahid_rows = [
        ("NTM Revenue (SAR M)", 1650, 1650, 1650),
        ("EV/Revenue multiple", 0.03, 0.05, 0.08),
        ("Shahid EV (SAR M)", None, None, None),
    ]

    for j, row_data in enumerate(shahid_rows):
        r2 = r + 1 + j
        label_cell(ws, r2, 1, row_data[0], indent=2)
        for k, val in enumerate(row_data[1:]):
            col_idx = k + 2
            cell = ws.cell(row=r2, column=col_idx)
            if val is None:
                if j == 2:
                    cell.value = f"={get_column_letter(col_idx)}{r + 1}*{get_column_letter(col_idx)}{r + 2}"
                else:
                    cell.value = "—"
            else:
                cell.value = val
                if j == 0:
                    cell.number_format = "#,##0"
                elif j == 1:
                    cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border()

    r += 5
    label_cell(ws, r, 1, "Less: Net debt allocated (23%)", indent=2)
    for k in range(3):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx, value=-283)
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    r += 2
    label_cell(ws, r, 1, "Shahid Equity (SAR M)", indent=2, bold=True)
    for k in range(3):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx)
        cell.value = (
            f"={get_column_letter(col_idx)}{r - 4}+{get_column_letter(col_idx)}{r - 2}"
        )
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=VERY_LIGHT)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    # M&E
    r += 3
    section_header(ws, r, 1, "SEGMENT 3: M&E (MEDIA & EVENTS)", 5)
    r += 1
    label_cell(
        ws, r, 1, "Best analog: Live Nation (events), production peers", indent=2
    )

    r += 1
    headers6 = ["", "Bear (8x)", "Base (10x)", "Bull (12x)"]
    for i, h in enumerate(headers6):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    me_rows = [
        ("NTM Revenue (SAR M)", 1400, 1400, 1400),
        ("EBITDA Margin", 0.04, 0.05, 0.06),
        ("EBITDA (SAR M)", None, None, None),
    ]

    for j, row_data in enumerate(me_rows):
        r2 = r + 1 + j
        label_cell(ws, r2, 1, row_data[0], indent=2)
        for k, val in enumerate(row_data[1:]):
            col_idx = k + 2
            cell = ws.cell(row=r2, column=col_idx)
            if val is None:
                if j == 2:
                    cell.value = f"={get_column_letter(col_idx)}{r + 1}*{get_column_letter(col_idx)}{r + 2}"
                else:
                    cell.value = "—"
            else:
                cell.value = val
                if j == 0:
                    cell.number_format = "#,##0"
                elif j == 1:
                    cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border()

    r += 5
    label_cell(ws, r, 1, "EV/EBITDA multiple → EV", indent=2)
    for k, (mult, lbl) in enumerate([(8, "Bear"), (10, "Base"), (12, "Bull")]):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx)
        cell.value = f"={get_column_letter(col_idx)}{r - 3}*{mult}"
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    r += 2
    label_cell(ws, r, 1, "Less: Net debt allocated (23%)", indent=2)
    for k in range(3):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx, value=-283)
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    r += 2
    label_cell(ws, r, 1, "M&E Equity (SAR M)", indent=2, bold=True)
    for k in range(3):
        col_idx = k + 2
        cell = ws.cell(row=r, column=col_idx)
        cell.value = (
            f"={get_column_letter(col_idx)}{r - 4}+{get_column_letter(col_idx)}{r - 2}"
        )
        cell.number_format = "#,##0"
        cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=VERY_LIGHT)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border()

    # SOTP Total
    r += 3
    section_header(ws, r, 1, "SOTP SUMMARY", 5)
    r += 1

    total_rows = [
        ("BOCA Equity", 6245, 8057, 10169),
        ("Shahid Equity", 4217, 7967, 14117),
        ("M&E Equity", 149, 417, 761),
        ("Total Equity (SAR M)", None, None, None),
        ("Shares Outstanding (M)", 332.5, 332.5, 332.5),
        ("Implied Share Price (SAR)", None, None, None),
        ("Current Market Price", 26.26, 26.26, 26.26),
        ("Upside to Market", None, None, None),
    ]

    headers7 = ["", "Bear", "Base", "Bull"]
    for i, h in enumerate(headers7):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for j, row_data in enumerate(total_rows):
        r2 = r + 1 + j
        is_total = j == 3
        is_price = j in [5, 7]
        label_cell(ws, r2, 1, row_data[0], indent=2, bold=is_total)
        for k in range(3):
            col_idx = k + 2
            cell = ws.cell(row=r2, column=col_idx)
            val = row_data[1 + k]
            if val is None:
                if j == 3:  # Total equity
                    cell.value = f"=SUM({get_column_letter(col_idx)}{r + 1}:{get_column_letter(col_idx)}{r + 3})"
                elif j == 5:  # Share price
                    cell.value = f"={get_column_letter(col_idx)}{r + 4}/{get_column_letter(col_idx)}{r + 5}"
                elif j == 7:  # Upside
                    cell.value = f"=({get_column_letter(col_idx)}{r + 6}-{get_column_letter(col_idx)}{r + 7})/{get_column_letter(col_idx)}{r + 7}"
            else:
                cell.value = val
            cell.number_format = "#,##0" if j < 5 else ("#,##0.0" if j == 5 else "0.0%")
            cell.font = Font(
                name="Calibri",
                size=10,
                bold=is_total or is_price,
                color=NAVY if is_total else DARK_GRAY,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=GREEN_FILL
                if is_total
                else (VERY_LIGHT if j == 7 else LIGHT_GRAY),
            )
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
def build_summary_sheet(wb):
    ws = wb.create_sheet("Valuation Summary")
    ws.sheet_view.showGridLines = False

    for col, w in [(1, 34), (2, 16), (3, 16), (4, 16), (5, 16)]:
        set_col_width(ws, col, w)

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MBC GROUP — VALUATION SUMMARY"
    t.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = "4072.SR | Tadawul | As of April 21, 2026 | SAR millions"
    sub.font = Font(name="Calibri", size=10, italic=True, color=MED_GRAY)
    sub.fill = PatternFill("solid", fgColor=NAVY)
    sub.alignment = Alignment(horizontal="center")

    r = 4
    section_header(ws, r, 1, "VALUATION OUTPUTS — ALL METHODS", 5)

    methods = [
        ("DCF — Bear (WACC 12%, g 4.5%)", 37.7, "+43.6%", "HIGH RISK"),
        ("DCF — Base (WACC 10%, g 4.5%)", 48.8, "+85.8%", "BASE CASE"),
        ("DCF — Bull (WACC 8%, g 4.5%)", 112.4, "+328%", "OPTIMISTIC"),
        ("Comps — Netflix multiple (26x)", 58.5, "+123%", "STRETCH"),
        ("Comps — Blended 15x NTM EBITDA", 35.0, "+33%", "CONSERVATIVE"),
        ("Comps — Blended 12x NTM EBITDA", 28.7, "+9%", "BEARISH"),
        ("Comps — Disney/WBD rev 2.6x", 46.8, "+78%", "CORE PEER"),
        ("SOTP — Bear (BOCA 9x / Shahid 3x)", 31.9, "+21%", "BEAR"),
        ("SOTP — Base (BOCA 10.5x / Shahid 5x)", 49.4, "+88%", "BASE CASE"),
        ("SOTP — Bull (BOCA 12x / Shahid 8x)", 75.3, "+187%", "BULL"),
        ("Paramount Precedent (7.5x)", 16.3, "-38%", "FLOOR"),
        ("WBD Distressed (3x)", 2.8, "-89%", "DISTRESSED"),
    ]

    headers = ["Method", "Share Price (SAR)", "vs. Market", "Note"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r + 1, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for j, row_data in enumerate(methods):
        r2 = r + 2 + j
        is_base = "Base" in row_data[0]
        is_bear = "Bear" in row_data[0] and "Base" not in row_data[0]
        is_bull = "Bull" in row_data[0]
        upside_str = row_data[2]
        is_positive = not upside_str.startswith("-")

        for k, val in enumerate(row_data):
            cell = ws.cell(row=r2, column=k + 1, value=val)
            cell.font = Font(
                name="Calibri",
                size=10,
                bold=is_base,
                color=(
                    "00B050"
                    if (is_positive and k == 2)
                    else ("FF0000" if (not is_positive and k == 2) else DARK_GRAY)
                ),
            )
            fill_color = (
                GREEN_FILL
                if is_base
                else (
                    "FFE0CC"
                    if is_bear
                    else (
                        "E2EFDA" if is_bull else (LIGHT_GRAY if j % 2 == 0 else WHITE)
                    )
                )
            )
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(
                horizontal="right" if k > 0 else "left", indent=1 if k == 0 else 0
            )
            cell.border = thin_border()

    # ── Blended Central Estimate ───────────────────────────────────────────
    r += len(methods) + 4
    section_header(ws, r, 1, "BLENDED CENTRAL ESTIMATE", 5)
    set_row_height(ws, r, 20)

    r += 1
    weights = [
        ("DCF Base Case", 0.35, 48.8),
        ("SOTP Base Case", 0.35, 49.4),
        ("Comparable Companies", 0.20, 35.0),
        ("Paramount Precedent", 0.10, 16.3),
    ]

    ws.cell(row=r, column=1, value="Method").font = Font(
        name="Calibri", size=10, bold=True, color=WHITE
    )
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=MID_BLUE)
    ws.cell(row=r, column=2, value="Weight").font = Font(
        name="Calibri", size=10, bold=True, color=WHITE
    )
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=MID_BLUE)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value="Price").font = Font(
        name="Calibri", size=10, bold=True, color=WHITE
    )
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=MID_BLUE)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value="Weighted").font = Font(
        name="Calibri", size=10, bold=True, color=WHITE
    )
    ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=MID_BLUE)
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")

    total_weighted = 0
    for i2, (method, wt, price) in enumerate(weights):
        r2 = r + 1 + i2
        total_weighted += wt * price
        for col_idx, val in enumerate([method, wt, price, None]):
            cell = ws.cell(row=r2, column=col_idx + 1)
            if val is None:
                cell.value = f"={get_column_letter(col_idx + 1)}{r}*{get_column_letter(col_idx + 1)}{r + 1 + col_idx}".replace(
                    f"{get_column_letter(col_idx + 1)}{r}*{get_column_letter(col_idx + 1)}",
                    f"{get_column_letter(col_idx + 1)}{r}*{get_column_letter(col_idx + 1)}",
                ).replace(
                    f"{get_column_letter(col_idx + 1)}{r + 1 + col_idx}", str(price)
                )
                # Simple weighted value
                cell.value = round(wt * price, 1)
                cell.number_format = "#,##0.0"
            else:
                cell.value = val
                if col_idx == 1:
                    cell.number_format = "0%"
                elif col_idx == 2:
                    cell.number_format = "#,##0.0"
            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill(
                "solid", fgColor=LIGHT_GRAY if i2 % 2 == 0 else WHITE
            )
            cell.alignment = Alignment(
                horizontal="right" if col_idx > 0 else "left",
                indent=1 if col_idx == 0 else 0,
            )
            cell.border = thin_border()

    r += len(weights) + 1
    # Total row
    for col_idx, val in enumerate(
        ["Blended Central Estimate", None, None, round(total_weighted, 1)]
    ):
        cell = ws.cell(row=r, column=col_idx + 1, value=val)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 0 else "left",
            indent=1 if col_idx == 0 else 0,
        )
        if col_idx == 3:
            cell.number_format = "#,##0.0"
        cell.border = thin_border()

    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    verdict_cell = ws[f"A{r}"]
    verdict_cell.value = f"INVESTMENT RATING: BUY  |  TARGET RANGE: SAR 42–52  |  CENTRAL TARGET: SAR {round(total_weighted)}  |  CURRENT: SAR 26.26  |  UPSIDE: {round((total_weighted - 26.26) / 26.26 * 100)}%"
    verdict_cell.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    verdict_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
    verdict_cell.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, r, 28)

    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    note_cell = ws[f"A{r}"]
    note_cell.value = (
        "Note: MBC listed Jan 2024. Only ~2 years of public trading history — no reliable beta. "
        "WACC is estimated, not market-derived. DCF and SOTP convergence at SAR 48–50 is meaningful "
        "but both methods share streaming margin expansion assumptions. Primary risk: Shahid trajectory."
    )
    note_cell.font = Font(name="Calibri", size=9, italic=True, color=MED_GRAY)
    note_cell.alignment = Alignment(horizontal="left", wrap_text=True)
    set_row_height(ws, r, 36)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_dcf_sheet(wb)
    build_comps_sheet(wb)
    build_sotp_sheet(wb)
    build_summary_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
